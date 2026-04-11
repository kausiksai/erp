"""Shared helpers for the Excel parsers.

The source files from the srimukha ERP share several quirks that this module
papers over in one place:

    * Files arrive with a `.xls` extension but are actually `.xlsx` content.
      openpyxl rejects them by filename, so we always read bytes and hand
      them to openpyxl through a BytesIO wrapper.

    * `read_only=True` sometimes reports bogus dimensions (1x1) for these
      files, so we read in the default mode. The tradeoff is higher memory
      for very large files; the largest we have seen is ASN at ~84k rows
      and that stays under ~300 MB RSS.

    * Date columns arrive in multiple formats: native datetime objects, ISO
      strings, DD-MM-YYYY, DD.MM.YYYY, DD/MM/YYYY. `coerce_date` tries each
      in priority order and returns `None` for unrecognisable values rather
      than raising, so a single bad cell does not poison the whole file.

    * Numeric columns are sometimes strings with trailing '.0' (e.g. the
      AMD_NO column comes through as '0.0'). `coerce_int` and `coerce_decimal`
      handle that.

The module is deliberately stateless; no shared global resources.
"""

from __future__ import annotations

import logging
import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple, Union

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

log = logging.getLogger(__name__)

PathOrBytes = Union[str, Path, bytes, BytesIO]

# ----------------------------------------------------------------------------
# Workbook loading
# ----------------------------------------------------------------------------
def load_workbook_flex(source: PathOrBytes) -> Any:
    """Load an Excel workbook from disk or bytes.

    Accepts a path (str / Path), raw bytes, or an existing BytesIO. Always
    routes through BytesIO to bypass openpyxl's filename-extension check,
    which rejects the legacy `.xls` name the srimukha ERP uses.
    """
    if isinstance(source, (str, Path)):
        p = Path(source)
        if not p.is_file():
            raise FileNotFoundError(f"Excel file not found: {p}")
        with p.open("rb") as fh:
            buf = BytesIO(fh.read())
    elif isinstance(source, bytes):
        buf = BytesIO(source)
    elif isinstance(source, BytesIO):
        buf = source
        buf.seek(0)
    else:
        raise TypeError(f"Unsupported source type: {type(source).__name__}")

    try:
        return load_workbook(buf, data_only=True)
    except Exception as exc:
        raise ParseError(f"openpyxl failed to open workbook: {exc}") from exc


class ParseError(RuntimeError):
    """Raised when a parser cannot make sense of an input file."""


# ----------------------------------------------------------------------------
# Header detection
# ----------------------------------------------------------------------------
def _norm(s: Any) -> str:
    """Canonicalise a header string for matching: lower-case, trim, collapse
    whitespace, strip trailing punctuation."""
    if s is None:
        return ""
    text = str(s).strip().lower()
    text = re.sub(r"\s+", " ", text)
    text = text.rstrip(" .:")
    return text


def build_header_map(
    ws: Worksheet,
    aliases: Dict[str, Sequence[str]],
    *,
    max_scan_rows: int = 5,
    required: Optional[Iterable[str]] = None,
) -> Tuple[int, Dict[str, int]]:
    """Locate the header row in `ws` and return `(header_row_index, col_map)`.

    `aliases` maps a canonical field name (e.g. 'po_number') to the list of
    accepted header variants for that field. The first variant that appears
    anywhere in the first `max_scan_rows` rows wins.

    Returns `(1-based header row number, {field_name: 0-based column index})`.

    If `required` is provided, raises ParseError on any missing field.
    """
    if ws.max_row == 0 or ws.max_column == 0:
        raise ParseError("Worksheet is empty")

    # Build lookup of normalised alias -> canonical field name
    alias_lookup: Dict[str, str] = {}
    for field, variants in aliases.items():
        for variant in variants:
            alias_lookup[_norm(variant)] = field

    best_row = 0
    best_map: Dict[str, int] = {}
    scan_limit = min(max_scan_rows, ws.max_row)

    for row_idx in range(1, scan_limit + 1):
        row_cells = [
            _norm(ws.cell(row=row_idx, column=col_idx).value)
            for col_idx in range(1, ws.max_column + 1)
        ]
        col_map: Dict[str, int] = {}
        for col_idx, cell in enumerate(row_cells):
            if cell and cell in alias_lookup:
                field = alias_lookup[cell]
                # first occurrence wins so that duplicate headers
                # (GRN has three columns literally named "Type") do not
                # overwrite an earlier match.
                col_map.setdefault(field, col_idx)
        if len(col_map) > len(best_map):
            best_map = col_map
            best_row = row_idx

    if best_row == 0:
        raise ParseError(
            f"Could not locate header row (scanned {scan_limit} rows). "
            f"Expected any of: {sorted(aliases.keys())}"
        )

    if required:
        missing = [f for f in required if f not in best_map]
        if missing:
            # Include the row we thought was the header for debugging.
            header_preview = [
                ws.cell(row=best_row, column=c).value
                for c in range(1, min(ws.max_column + 1, 40))
            ]
            raise ParseError(
                f"Header row {best_row} is missing required fields {missing}. "
                f"Row preview: {header_preview!r}"
            )

    return best_row, best_map


# ----------------------------------------------------------------------------
# Value coercion
# ----------------------------------------------------------------------------
def coerce_str(value: Any, *, max_len: Optional[int] = None) -> Optional[str]:
    """Trim, collapse whitespace, return None for empty."""
    if value is None:
        return None
    if isinstance(value, (int, float, Decimal)):
        text = str(value)
    else:
        text = str(value).strip()
    if not text:
        return None
    text = re.sub(r"\s+", " ", text)
    if max_len is not None and len(text) > max_len:
        text = text[:max_len]
    return text


_DATE_FORMATS = (
    "%Y-%m-%d %H:%M:%S",
    "%Y-%m-%d",
    "%Y/%m/%d",
    "%d-%m-%Y %H:%M:%S",
    "%d-%m-%Y",
    "%d/%m/%Y",
    "%d.%m.%Y",
    "%m/%d/%Y",   # last resort; source files do not use this, but the OCR
                  # pipeline sometimes does — harmless to keep as a fallback
)


def coerce_date(value: Any) -> Optional[date]:
    """Return a Python `date` or None. Never raises."""
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        return None
    # Strip trailing time chunks that may sneak in from Excel general cells
    text = text.replace("T", " ")
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    log.debug("coerce_date: unrecognised %r", value)
    return None


def coerce_int(value: Any) -> Optional[int]:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        # Reject NaN / inf
        if value != value or value in (float("inf"), float("-inf")):
            return None
        return int(value)
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    try:
        return int(text)
    except ValueError:
        pass
    try:
        return int(float(text))
    except (ValueError, OverflowError):
        log.debug("coerce_int: unrecognised %r", value)
        return None


def coerce_decimal(value: Any) -> Optional[Decimal]:
    if value is None or value == "":
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, bool):
        return Decimal(int(value))
    if isinstance(value, (int, float)):
        if isinstance(value, float) and (value != value or value in (float("inf"), float("-inf"))):
            return None
        try:
            return Decimal(str(value))
        except (InvalidOperation, ValueError):
            return None
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    try:
        return Decimal(text)
    except (InvalidOperation, ValueError):
        log.debug("coerce_decimal: unrecognised %r", value)
        return None


def coerce_bool(value: Any) -> Optional[bool]:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return value
    text = str(value).strip().lower()
    if text in {"y", "yes", "true", "1", "t"}:
        return True
    if text in {"n", "no", "false", "0", "f"}:
        return False
    return None


# ----------------------------------------------------------------------------
# Row iteration
# ----------------------------------------------------------------------------
def iter_data_rows(
    ws: Worksheet,
    header_row: int,
    col_map: Dict[str, int],
) -> Iterable[Dict[str, Any]]:
    """Yield {field: raw_value} dicts for every non-empty data row.

    Caller is responsible for coercing values. This helper only strips
    fully-empty rows so that blank trailing rows in the workbook do not
    appear in the output.
    """
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if row is None:
            continue
        if all(c is None or (isinstance(c, str) and c.strip() == "") for c in row):
            continue
        out: Dict[str, Any] = {}
        for field, col_idx in col_map.items():
            if col_idx < len(row):
                out[field] = row[col_idx]
            else:
                out[field] = None
        yield out


def select_sheet(wb: Any, preferred_names: Sequence[str] = ()) -> Worksheet:
    """Return the first matching sheet by name, else the active sheet.

    srimukha exports always ship a single sheet called 'NewSheet1', but
    parsers accept a preferred-name hint so future variations are easy.
    """
    for name in preferred_names:
        if name in wb.sheetnames:
            return wb[name]
    return wb[wb.sheetnames[0]]
